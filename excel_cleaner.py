import os
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl import load_workbook

os.environ['TK_SILENCE_DEPRECATION'] = '1'

class ExcelColumnClearerGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel 列清空工具")
        self.root.geometry("400x300")  # 调整窗口大小

        self.folder_path = ""
        self.column_number = 0  # 处理的列号
        self.create_widgets()

    def create_widgets(self):
        # 文件夹选择按钮
        btn_folder = tk.Button(self.root, text="选择文件夹", command=self.select_folder)
        btn_folder.pack(pady=10)

        self.folder_label = tk.Label(self.root, text="未选择文件夹")
        self.folder_label.pack()

        # 列号输入部分
        lbl_column = tk.Label(self.root, text="输入列号（如：3）：")
        lbl_column.pack(pady=10)

        # 自定义输入框颜色和背景色
        self.column_entry = tk.Entry(self.root, width=20, bg="#e0e0e0", fg="black")  # 背景色和前景色
        self.column_entry.pack()

        # 执行按钮
        btn_start = tk.Button(self.root, text="开始处理", command=self.process_files)
        btn_start.pack(pady=20)

    def select_folder(self):
        folder_path = filedialog.askdirectory(title="选择目标文件夹")
        if folder_path:
            self.folder_path = folder_path
            self.folder_label.config(text=f"已选择：{os.path.basename(folder_path)}")

    def process_files(self):
        # 获取用户输入的列号
        try:
            self.column_number = int(self.column_entry.get().strip())
            if self.column_number <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("错误", "请输入有效的列号！")
            return

        if not self.folder_path:
            messagebox.showerror("错误", "请先选择目标文件夹！")
            return

        try:
            updated_count = self.clear_column_in_files()
            messagebox.showinfo("完成", f"成功处理 {updated_count} 个Excel文件！")
        except Exception as e:
            messagebox.showerror("错误", f"处理过程中出错：{str(e)}")

    def clear_column_in_files(self):
        processed_files = 0

        for root, dirs, files in os.walk(self.folder_path):
            for file in files:
                if file.endswith(('.xlsx', '.xls')):
                    file_path = os.path.join(root, file)
                    try:
                        # 使用 openpyxl 加载工作簿以保留格式
                        wb = load_workbook(file_path)
                        ws = wb.active
                        
                        # 获取要清空的列的字母索引
                        from openpyxl.utils import get_column_letter
                        col_letter = get_column_letter(self.column_number)
                        
                        # 清空指定列的内容（跳过表头）
                        for row in range(2, ws.max_row + 1):
                            ws[f'{col_letter}{row}'].value = None
                            
                        # 保存文件
                        wb.save(file_path)
                        processed_files += 1

                    except Exception as e:
                        print(f"处理文件 {file} 时出错：{str(e)}")
                        continue

        return processed_files

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = ExcelColumnClearerGUI()
    app.run()